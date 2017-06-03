from scrapy import FormRequest
from ..items import GetProfileItem
from scrapy.spiders import CrawlSpider
from openpyxl import load_workbook


class GetProfileSpider(CrawlSpider):
    name = "get_profile_url"

    allowed_domains = ["www.texasbar.com"]

    start_urls = [
        'https://www.texasbar.com/AM/Template.cfm?Section=Find_A_Lawyer&Template=/CustomSource/MemberDirectory/Search_Form_Client_Main.cfm']

    def get_person(self):
        wb = load_workbook(filename='Names.xlsx')
        name_sheet = wb.get_sheet_names()[0]
        sheet = wb[name_sheet]
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]

    def parse(self, response):
        for each in list(self.get_person())[1:]:
            yield FormRequest.from_response(response,
                                            formnumber=1,
                                            formdata={'FirstName': each[1], 'LastName': each[0]},
                                            callback=self.parse_next)

    def parse_next(self, response):
        if b'<p>Your search has returned no result.</p>' not in response.body:
            item = GetProfileItem()
            item['url_profile'] = 'https://%s%s' % (self.allowed_domains[0],
                                                    response.selector.css('a.read-more::attr(href)').extract_first())
            return item
